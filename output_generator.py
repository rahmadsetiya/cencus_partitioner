"""
output_generator.py
===================
Menghasilkan laporan Excel dan ringkasan terminal dari hasil partisi.

Sheet yang dihasilkan:
1. "Ringkasan"   → statistik per petugas
2. "Detail SLS"  → daftar semua SLS dengan assignment petugas
3. "Adjacency"   → semua edge dalam graph (untuk referensi)
"""

import logging
from typing import Dict, List
import pandas as pd
import numpy as np
import networkx as nx
import geopandas as gpd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)

Partition = Dict[str, int]   # kode_sls → group_id

# Palet warna untuk tiap petugas (Excel hex, max 20 petugas)
GROUP_COLORS = [
    "4472C4", "ED7D31", "A9D18E", "FF0000", "FFFF00",
    "70AD47", "264478", "9E480E", "636363", "997300",
    "255E91", "843C0C", "622E44", "806000", "375623",
    "0563C1", "954F72", "C55A11", "538135", "833C00",
]


class OutputGenerator:
    """Membuat file Excel dan ringkasan hasil partisi."""

    def __init__(
        self,
        gdf: gpd.GeoDataFrame,
        G: nx.Graph,
        partition: Partition,
        n_groups: int,
    ):
        self.gdf       = gdf
        self.G         = G
        self.partition = partition
        self.n_groups  = n_groups

        # Pre-compute statistik
        self.group_stats = self._compute_group_stats()

    # =========================================================================
    # PUBLIC: Save Excel
    # =========================================================================

    def save_excel(self, filepath: str) -> None:
        """
        Simpan hasil partisi ke file Excel dengan formatting yang rapi.

        Parameters
        ----------
        filepath : str
            Path output file Excel (.xlsx).
        """
        # Buat DataFrames untuk tiap sheet
        df_summary = self._build_summary_df()
        df_detail  = self._build_detail_df()
        df_edges   = self._build_edges_df()

        # Tulis ke Excel dengan openpyxl engine
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Ringkasan", index=False)
            df_detail.to_excel(writer, sheet_name="Detail SLS", index=False)
            df_edges.to_excel(writer, sheet_name="Adjacency Graph", index=False)

        # Apply formatting post-processing
        self._format_excel(filepath, df_summary, df_detail)

        logger.info(f"  Output Excel tersimpan: {filepath}")

    # =========================================================================
    # PUBLIC: Terminal summary
    # =========================================================================

    def print_summary(self) -> None:
        """Cetak ringkasan hasil ke terminal."""
        print("\n" + "=" * 65)
        print("  HASIL PARTISI WILAYAH SENSUS")
        print("=" * 65)
        print(f"  Total SLS     : {len(self.gdf):,}")
        print(f"  Total muatan  : {self.gdf[config.COL_MUATAN].sum():,.0f}")
        print(f"  Jumlah petugas: {self.n_groups}")
        print(f"  Target/petugas: {self.gdf[config.COL_MUATAN].sum() / self.n_groups:,.1f}")
        print("-" * 65)
        print(f"  {'Petugas':<10} {'Jml SLS':>8} {'Muatan':>10} {'Connected':>12} {'Travel Cost':>12}")
        print("-" * 65)

        for group_id in range(self.n_groups):
            stats = self.group_stats[group_id]
            status = "YA" if stats["is_connected"] else "TIDAK ⚠"
            print(
                f"  Petugas {group_id+1:<3} "
                f"{stats['n_sls']:>8,} "
                f"{stats['total_muatan']:>10,.0f} "
                f"{'':>4}{status:>9} "
                f"{stats['total_travel_cost']:>11.2f}"
            )

        print("-" * 65)
        loads = [self.group_stats[g]["total_muatan"] for g in range(self.n_groups)]
        print(f"  Muatan maks   : {max(loads):,.0f}")
        print(f"  Muatan min    : {min(loads):,.0f}")
        print(f"  Selisih maks  : {max(loads) - min(loads):,.0f}")
        print(f"  CV (std/mean) : {np.std(loads)/np.mean(loads):.4f}")
        print("=" * 65 + "\n")

    # =========================================================================
    # PRIVATE: Build DataFrames
    # =========================================================================

    def _build_summary_df(self) -> pd.DataFrame:
        """DataFrame ringkasan per petugas."""
        rows = []
        for group_id in range(self.n_groups):
            stats = self.group_stats[group_id]
            rows.append({
                "Petugas": f"Petugas {group_id + 1}",
                "Jumlah SLS": stats["n_sls"],
                "Total Muatan": stats["total_muatan"],
                "Muatan Min SLS": stats["min_sls_load"],
                "Muatan Maks SLS": stats["max_sls_load"],
                "Connected": "Ya" if stats["is_connected"] else "TIDAK",
                "Total Travel Cost": round(stats["total_travel_cost"], 2),
                "Daftar SLS": ", ".join(stats["sls_list"]),
            })
        return pd.DataFrame(rows)

    def _build_detail_df(self) -> pd.DataFrame:
        """DataFrame detail per SLS dengan assignment petugas."""
        rows = []
        for _, row in self.gdf.iterrows():
            kode = row[config.COL_KODE_SLS]
            group_id = self.partition.get(kode, -1)
            rows.append({
                "kode_sls": kode,
                "muatan": row[config.COL_MUATAN],
                "petugas": f"Petugas {group_id + 1}" if group_id >= 0 else "UNASSIGNED",
                "group_id": group_id + 1 if group_id >= 0 else -1,
                "centroid_lon": round(row.get("centroid_lon", 0), 6),
                "centroid_lat": round(row.get("centroid_lat", 0), 6),
            })

        df = pd.DataFrame(rows)
        # Urutkan berdasarkan petugas lalu kode_sls
        df = df.sort_values(["petugas", "kode_sls"]).reset_index(drop=True)
        return df

    def _build_edges_df(self) -> pd.DataFrame:
        """DataFrame semua edge di graph (untuk referensi dan audit)."""
        rows = []
        for u, v, data in self.G.edges(data=True):
            group_u = self.partition.get(u, -1)
            group_v = self.partition.get(v, -1)
            rows.append({
                "kode_sls_a": u,
                "kode_sls_b": v,
                "weight": round(data.get("weight", -1), 4),
                "road_dist_m": data.get("road_dist_m", -1),
                "is_touching": data.get("is_touching", False),
                "manual_override": data.get("manual_override", ""),
                "same_group": "Ya" if group_u == group_v else "Tidak",
                "petugas_a": f"Petugas {group_u + 1}" if group_u >= 0 else "?",
                "petugas_b": f"Petugas {group_v + 1}" if group_v >= 0 else "?",
            })
        return pd.DataFrame(rows)

    # =========================================================================
    # PRIVATE: Compute stats
    # =========================================================================

    def _compute_group_stats(self) -> Dict:
        """Hitung statistik untuk setiap grup."""
        stats = {}

        # Inisialisasi
        for g in range(self.n_groups):
            stats[g] = {
                "n_sls": 0,
                "total_muatan": 0.0,
                "min_sls_load": float("inf"),
                "max_sls_load": 0.0,
                "sls_list": [],
                "is_connected": True,
                "total_travel_cost": 0.0,
            }

        # Akumulasi muatan dari GDF
        for _, row in self.gdf.iterrows():
            kode = row[config.COL_KODE_SLS]
            group_id = self.partition.get(kode, -1)
            if group_id < 0:
                continue

            muatan = float(row[config.COL_MUATAN])
            stats[group_id]["n_sls"]          += 1
            stats[group_id]["total_muatan"]   += muatan
            stats[group_id]["sls_list"].append(kode)
            stats[group_id]["min_sls_load"] = min(
                stats[group_id]["min_sls_load"], muatan
            )
            stats[group_id]["max_sls_load"] = max(
                stats[group_id]["max_sls_load"], muatan
            )

        # Cek konektivitas dan hitung travel cost per grup
        for g in range(self.n_groups):
            sls_nodes = stats[g]["sls_list"]
            if not sls_nodes:
                continue

            # Fix min jika tidak ada node (edge case)
            if stats[g]["min_sls_load"] == float("inf"):
                stats[g]["min_sls_load"] = 0.0

            # Subgraf untuk cek konektivitas
            subgraph = self.G.subgraph(sls_nodes)
            stats[g]["is_connected"] = (
                len(sls_nodes) <= 1 or nx.is_connected(subgraph)
            )

            # Total travel cost = jumlah weight semua edge internal grup
            stats[g]["total_travel_cost"] = sum(
                data.get("weight", 0)
                for _, _, data in subgraph.edges(data=True)
            )

        return stats

    # =========================================================================
    # PRIVATE: Excel formatting
    # =========================================================================

    def _format_excel(
        self,
        filepath: str,
        df_summary: pd.DataFrame,
        df_detail: pd.DataFrame,
    ) -> None:
        """Apply formatting ke file Excel yang sudah dibuat."""
        try:
            wb = openpyxl.load_workbook(filepath)

            self._format_sheet_summary(wb["Ringkasan"], df_summary)
            self._format_sheet_detail(wb["Detail SLS"], df_detail)
            self._format_sheet_adjacency(wb["Adjacency Graph"])

            wb.save(filepath)

        except Exception as e:
            logger.warning(f"  Formatting Excel gagal: {e}. File tetap tersimpan tanpa format.")

    def _format_sheet_summary(self, ws, df: pd.DataFrame) -> None:
        """Format sheet Ringkasan."""
        # Header style
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Warnai tiap baris sesuai petugas
        for row_idx in range(2, len(df) + 2):
            group_id = row_idx - 2  # 0-indexed
            color = GROUP_COLORS[group_id % len(GROUP_COLORS)]
            row_fill = PatternFill("solid", fgColor=color + "33")  # transparan 20%

            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # Auto-fit lebar kolom
        _autofit_columns(ws)

    def _format_sheet_detail(self, ws, df: pd.DataFrame) -> None:
        """Format sheet Detail SLS dengan warna per petugas."""
        # Header
        header_fill = PatternFill("solid", fgColor="375623")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Cari kolom group_id
        try:
            group_col_idx = list(df.columns).index("group_id") + 1
        except ValueError:
            group_col_idx = None

        # Warnai baris berdasarkan grup
        for row_idx in range(2, len(df) + 2):
            if group_col_idx:
                group_id = ws.cell(row=row_idx, column=group_col_idx).value
                if group_id and isinstance(group_id, int) and group_id > 0:
                    color = GROUP_COLORS[(group_id - 1) % len(GROUP_COLORS)]
                    row_fill = PatternFill("solid", fgColor=color + "44")
                    for col_idx in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = row_fill

        _autofit_columns(ws)

    def _format_sheet_adjacency(self, ws) -> None:
        """Format sheet Adjacency Graph."""
        header_fill = PatternFill("solid", fgColor="7B7B7B")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Highlight edge yang berbeda grup (cross-group edges)
        try:
            same_group_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == "same_group":
                    same_group_col = col_idx
                    break

            if same_group_col:
                for row in ws.iter_rows(min_row=2):
                    val = row[same_group_col - 1].value
                    if val == "Tidak":
                        for cell in row:
                            cell.fill = PatternFill("solid", fgColor="FFE699")
        except Exception:
            pass

        _autofit_columns(ws)


# =============================================================================
# HELPER
# =============================================================================

def _autofit_columns(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit lebar kolom berdasarkan konten."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width
