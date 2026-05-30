"""
manual_override.py
==================
Membaca file Excel manual_override.xlsx dan menerapkan koreksi
adjacency berdasarkan pengetahuan lapangan petugas.

Format file Excel yang diharapkan:
Sheet "force_connect":
    | kode_sls_a | kode_sls_b | catatan       |
    |------------|------------|---------------|
    | SLS_001    | SLS_003    | akses jembatan|

Sheet "force_disconnect":
    | kode_sls_a | kode_sls_b | catatan          |
    |------------|------------|------------------|
    | SLS_007    | SLS_008    | sungai tanpa jembatan|
"""

import logging
from pathlib import Path

import networkx as nx
import pandas as pd

import config

logger = logging.getLogger(__name__)

# Nama sheet di file Excel
SHEET_CONNECT = "force_connect"
SHEET_DISCONNECT = "force_disconnect"

# Nama kolom wajib di tiap sheet
COL_A = "kode_sls_a"
COL_B = "kode_sls_b"
COL_NOTE = "catatan"


# =============================================================================
# PUBLIC FUNCTION
# =============================================================================


def apply_manual_override(G: nx.Graph, filepath: str) -> nx.Graph:
    """
    Terapkan manual override dari file Excel ke graf aksesibilitas.

    Parameters
    ----------
    G : nx.Graph
        Graf aksesibilitas yang sudah dibangun.
    filepath : str
        Path ke file manual_override.xlsx.

    Returns
    -------
    nx.Graph
        Graf yang sudah dimodifikasi sesuai override.
    """
    path = Path(filepath)
    if not path.exists():
        logger.warning(f"  File override tidak ditemukan: {filepath}")
        return G

    G = G.copy()

    # -------------------------------------------------------------------------
    # 1. Force Connect
    # -------------------------------------------------------------------------
    df_connect = _read_sheet(path, SHEET_CONNECT)
    if df_connect is not None:
        n_connected = 0
        for _, row in df_connect.iterrows():
            kode_a = str(row[COL_A]).strip()
            kode_b = str(row[COL_B]).strip()
            catatan = row.get(COL_NOTE, "")

            # Validasi: kedua node harus ada di graph
            if not G.has_node(kode_a):
                logger.warning(f"  [override] Node tidak ditemukan: {kode_a}")
                continue
            if not G.has_node(kode_b):
                logger.warning(f"  [override] Node tidak ditemukan: {kode_b}")
                continue

            if G.has_edge(kode_a, kode_b):
                logger.debug(f"  [override] Edge {kode_a}↔{kode_b} sudah ada, dilewati.")
                continue

            # Tambahkan edge dengan weight kecil (akses mudah via manual)
            G.add_edge(
                kode_a,
                kode_b,
                weight=config.DEFAULT_TOUCHING_WEIGHT,
                road_dist_m=-1,
                is_touching=False,
                manual_override="force_connect",
                catatan=str(catatan),
            )
            logger.info(
                f"  [FORCE CONNECT] {kode_a} ↔ {kode_b}" + (f" ({catatan})" if catatan else "")
            )
            n_connected += 1

        logger.info(f"  Force connect diterapkan: {n_connected} edge baru")

    # -------------------------------------------------------------------------
    # 2. Force Disconnect
    # -------------------------------------------------------------------------
    df_disconnect = _read_sheet(path, SHEET_DISCONNECT)
    if df_disconnect is not None:
        n_disconnected = 0
        for _, row in df_disconnect.iterrows():
            kode_a = str(row[COL_A]).strip()
            kode_b = str(row[COL_B]).strip()
            catatan = row.get(COL_NOTE, "")

            if not G.has_edge(kode_a, kode_b):
                logger.debug(
                    f"  [override] Edge {kode_a}↔{kode_b} tidak ada, tidak perlu disconnect."
                )
                continue

            G.remove_edge(kode_a, kode_b)
            logger.info(
                f"  [FORCE DISCONNECT] {kode_a} ↔ {kode_b}" + (f" ({catatan})" if catatan else "")
            )
            n_disconnected += 1

            # Cek apakah force disconnect membuat node terisolasi
            if G.degree(kode_a) == 0:
                logger.warning(
                    f"  PERINGATAN: {kode_a} menjadi isolated node setelah force disconnect!"
                )
            if G.degree(kode_b) == 0:
                logger.warning(
                    f"  PERINGATAN: {kode_b} menjadi isolated node setelah force disconnect!"
                )

        logger.info(f"  Force disconnect diterapkan: {n_disconnected} edge dihapus")

    # -------------------------------------------------------------------------
    # 3. Cek konektivitas setelah override
    # -------------------------------------------------------------------------
    if not nx.is_connected(G):
        n_comp = nx.number_connected_components(G)
        logger.warning(
            f"  PERINGATAN: Graf memiliki {n_comp} komponen terpisah "
            f"setelah override. Periksa force_disconnect."
        )

    return G


# =============================================================================
# PRIVATE HELPERS
# =============================================================================


def _read_sheet(
    path: Path,
    sheet_name: str,
) -> pd.DataFrame | None:
    """
    Baca satu sheet dari file Excel.

    Returns None jika sheet tidak ada atau kosong.
    """
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
        df.columns = [c.strip().lower() for c in df.columns]

        # Pastikan kolom wajib ada
        missing = [c for c in [COL_A, COL_B] if c not in df.columns]
        if missing:
            logger.warning(
                f"  Sheet '{sheet_name}' tidak memiliki kolom: {missing}. "
                f"Kolom tersedia: {list(df.columns)}"
            )
            return None

        # Hapus baris yang kosong
        df = df.dropna(subset=[COL_A, COL_B])

        if df.empty:
            logger.info(f"  Sheet '{sheet_name}' kosong.")
            return None

        # Tambahkan kolom catatan jika tidak ada
        if COL_NOTE not in df.columns:
            df[COL_NOTE] = ""

        logger.info(f"  Sheet '{sheet_name}': {len(df)} baris override ditemukan.")
        return df

    except Exception as e:
        logger.warning(f"  Gagal membaca sheet '{sheet_name}': {e}")
        return None


# =============================================================================
# UTILITY: Template generator
# =============================================================================


def generate_override_template(output_path: str = "manual_override_template.xlsx") -> None:
    """
    Generate template file manual_override.xlsx yang kosong.

    Berguna untuk distribusi ke tim lapangan agar mereka tahu
    format yang benar.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # Hapus default sheet
    wb.remove(wb.active)

    header_fill_connect = PatternFill("solid", fgColor="1E7B34")
    header_fill_disconnect = PatternFill("solid", fgColor="C0392B")

    for sheet_name, fill_color in [
        (SHEET_CONNECT, header_fill_connect),
        (SHEET_DISCONNECT, header_fill_disconnect),
    ]:
        ws = wb.create_sheet(sheet_name)

        # Header
        headers = [COL_A, COL_B, COL_NOTE]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill_color
            cell.alignment = Alignment(horizontal="center")

        # Contoh baris
        if sheet_name == SHEET_CONNECT:
            ws.append(["SLS_001", "SLS_003", "Akses via jembatan desa"])
        else:
            ws.append(["SLS_007", "SLS_008", "Sungai tanpa jembatan, akses via SLS_010"])

        # Lebar kolom
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 40

    wb.save(output_path)
    logger.info(f"  Template override tersimpan: {output_path}")


if __name__ == "__main__":
    generate_override_template()
    print("Template berhasil dibuat: manual_override_template.xlsx")
